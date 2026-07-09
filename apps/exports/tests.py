from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

from django.contrib import admin
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from apps.counterparties.models import BudgetItem, Category, ChartOfAccount, Counterparty, CostCenter, Work
from apps.payments.models import Payment

from .models import ExportBatch
from .selectors import approved_payments_for_export
from .services import ACCOUNTING_HEADERS, ACCOUNTING_SHEET_NAME, HEADERS, SHEET_NAME, export_approved_payments


class ExportBatchTests(TestCase):
    def test_export_batch_is_admin_registered(self):
        self.assertTrue(admin.site.is_registered(ExportBatch))

    def test_export_batch_groups_approved_payments(self):
        payment = Payment.objects.create(amount=Decimal("123.45"), status=Payment.Status.APPROVED)
        batch = ExportBatch.objects.create(
            status=ExportBatch.Status.GENERATED,
            template_path="files/planilhas_modelo_importacao/Planilha_Modelo_Pagamentos.xlsx",
            records_count=1,
        )
        batch.payments.add(payment)

        self.assertEqual(batch.payments.count(), 1)
        self.assertEqual(payment.export_batches.first(), batch)

    def test_export_candidate_queryset_should_only_use_approved_payments(self):
        approved = Payment.objects.create(amount=Decimal("123.45"), status=Payment.Status.APPROVED)
        Payment.objects.create(amount=Decimal("55.00"), status=Payment.Status.RECEIVED)
        Payment.objects.create(amount=Decimal("77.00"), status=Payment.Status.CANCELED)

        candidates = approved_payments_for_export()

        self.assertEqual(list(candidates), [approved])


class PaymentSpreadsheetExportTests(TestCase):
    def setUp(self):
        self.tempdir = TemporaryDirectory()
        self.override = override_settings(
            MEDIA_ROOT=Path(self.tempdir.name),
            DEFAULT_PAYER="Company",
            DEFAULT_BANK_ACCOUNT="Banco Principal",
        )
        self.override.enable()
        self.category = Category.objects.create(name="Materiais", normalized_name="materiais")
        self.chart_account = ChartOfAccount.objects.create(name="Materiais", normalized_name="materiais")
        self.cost_center = CostCenter.objects.create(name="Project", normalized_name="project")
        self.work = Work.objects.create(name="Sertãozinho", normalized_name="sertaozinho")
        self.budget_item = BudgetItem.objects.create(
            work=self.work,
            index="3.4",
            item_type=BudgetItem.ItemType.SUBSTAGE,
            description="CALÇADA",
            normalized_description="calcada",
        )
        self.counterparty = Counterparty.objects.create(
            name="ACME Materiais",
            normalized_name="acme materiais",
            kind=Counterparty.Kind.SUPPLIER,
            primary_document="12345678000199",
        )

    def tearDown(self):
        self.override.disable()
        self.tempdir.cleanup()

    def test_only_approved_payments_are_exported(self):
        approved = self.make_payment(status=Payment.Status.APPROVED, document_number="NF-1")
        self.make_payment(status=Payment.Status.RECEIVED, document_number="NF-2")
        self.make_payment(status=Payment.Status.CANCELED, document_number="NF-3")

        batch = export_approved_payments()
        worksheet = self.load_export_sheet(batch)

        self.assertEqual(batch.records_count, 1)
        self.assertEqual(list(batch.payments.all()), [approved])
        self.assertEqual(worksheet.cell(row=2, column=7).value, "NF-1")
        self.assertIsNone(worksheet.cell(row=3, column=7).value)

    def test_required_columns_are_filled_in_template_order(self):
        self.make_payment(
            status=Payment.Status.APPROVED,
            amount=Decimal("123.45"),
            description="Compra de materiais",
            document_number="NF-123",
            payment_method="PIX",
            work_item_index="3.4",
        )

        batch = export_approved_payments()
        worksheet = self.load_export_sheet(batch)

        headers = [worksheet.cell(row=1, column=column).value for column in range(1, len(HEADERS) + 1)]
        self.assertEqual(headers, HEADERS)
        self.assertEqual(worksheet.cell(row=2, column=1).value.date(), date(2026, 6, 1))
        self.assertEqual(worksheet.cell(row=2, column=2).value.date(), date(2026, 6, 5))
        self.assertEqual(worksheet.cell(row=2, column=3).value.date(), date(2026, 6, 6))
        self.assertEqual(worksheet.cell(row=2, column=4).value, 123.45)
        self.assertEqual(worksheet.cell(row=2, column=5).value, "ACME Materiais")
        self.assertEqual(worksheet.cell(row=2, column=6).value, "Compra de materiais")
        self.assertEqual(worksheet.cell(row=2, column=7).value, "NF-123")
        self.assertEqual(worksheet.cell(row=2, column=8).value, "Materiais")
        self.assertEqual(worksheet.cell(row=2, column=9).value, "PIX")
        self.assertEqual(worksheet.cell(row=2, column=10).value, "Company")
        self.assertEqual(worksheet.cell(row=2, column=11).value, "Banco Principal")
        self.assertEqual(worksheet.cell(row=2, column=12).value, "Project")
        self.assertEqual(worksheet.cell(row=2, column=13).value, "Sertãozinho")
        self.assertEqual(worksheet.cell(row=2, column=14).value, "3.4")

    def test_description_is_limited_to_200_characters(self):
        self.make_payment(status=Payment.Status.APPROVED, description="x" * 260)

        batch = export_approved_payments()
        worksheet = self.load_export_sheet(batch)

        self.assertEqual(len(worksheet.cell(row=2, column=6).value), 200)

    def test_amount_uses_two_decimal_places(self):
        self.make_payment(status=Payment.Status.APPROVED, amount=Decimal("123.4"))

        batch = export_approved_payments()
        worksheet = self.load_export_sheet(batch)

        self.assertEqual(worksheet.cell(row=2, column=4).value, 123.4)
        self.assertIn("0.00", worksheet.cell(row=2, column=4).number_format)

    def test_export_batch_registers_exported_payments(self):
        first = self.make_payment(status=Payment.Status.APPROVED, document_number="NF-1")
        second = self.make_payment(status=Payment.Status.APPROVED, document_number="NF-2")

        batch = export_approved_payments()

        self.assertEqual(batch.status, ExportBatch.Status.GENERATED)
        self.assertEqual(batch.records_count, 2)
        self.assertTrue(batch.file.name.endswith(".xlsx"))
        self.assertTrue(batch.import_file.name.endswith(".xlsx"))
        self.assertTrue(batch.accounting_file.name.endswith(".xlsx"))
        self.assertIsNotNone(batch.generated_at)
        self.assertEqual(set(batch.payments.all()), {first, second})

    def test_accounting_spreadsheet_is_generated_for_accountant_model(self):
        self.make_payment(
            status=Payment.Status.APPROVED,
            amount=Decimal("123.45"),
            description="Compra de materiais",
            document_number="NF-123",
            payment_method="PIX",
            work_item_index="3.4",
        )

        batch = export_approved_payments()
        worksheet = self.load_accounting_sheet(batch)

        headers = [worksheet.cell(row=1, column=column).value for column in range(1, len(ACCOUNTING_HEADERS) + 1)]
        self.assertEqual(headers, ACCOUNTING_HEADERS)
        self.assertEqual(worksheet.cell(row=2, column=1).value, "1")
        self.assertEqual(worksheet.cell(row=2, column=5).value, 123.45)
        self.assertEqual(worksheet.cell(row=2, column=6).value, 0)
        self.assertEqual(worksheet.cell(row=2, column=10).value, 123.45)
        self.assertEqual(worksheet.cell(row=2, column=11).value, "ACME Materiais")
        self.assertEqual(worksheet.cell(row=2, column=12).value, "12345678000199")
        self.assertEqual(worksheet.cell(row=2, column=16).value, "Materiais")
        self.assertEqual(worksheet.cell(row=2, column=17).value, "Materiais")
        self.assertEqual(worksheet.cell(row=2, column=18).value, "Materiais")
        self.assertEqual(worksheet.cell(row=2, column=19).value, "À Vista")
        self.assertEqual(worksheet.cell(row=2, column=23).value, "Project")
        self.assertEqual(worksheet.cell(row=2, column=24).value, "Sertãozinho")
        self.assertEqual(worksheet.cell(row=2, column=25).value, "3.4")
        self.assertEqual(worksheet.cell(row=2, column=26).value, "CALÇADA")

    def test_export_does_not_fail_for_work_without_budget_item(self):
        work_without_budget = Work.objects.create(name="Tacima", normalized_name="tacima")
        self.make_payment(
            status=Payment.Status.APPROVED,
            work=work_without_budget,
            work_item_index="",
        )

        batch = export_approved_payments()
        import_sheet = self.load_export_sheet(batch)
        accounting_sheet = self.load_accounting_sheet(batch)

        self.assertEqual(batch.records_count, 1)
        self.assertEqual(import_sheet.cell(row=2, column=13).value, "Tacima")
        self.assertIn(import_sheet.cell(row=2, column=14).value, {None, ""})
        self.assertEqual(accounting_sheet.cell(row=2, column=24).value, "Tacima")
        self.assertEqual(accounting_sheet.cell(row=2, column=25).value, "-")
        self.assertEqual(accounting_sheet.cell(row=2, column=26).value, "-")

    def make_payment(
        self,
        status,
        amount=Decimal("100.00"),
        description="Payment de teste",
        document_number="NF-1",
        payment_method="PIX",
        work_item_index="3.4",
        work=None,
    ):
        return Payment.objects.create(
            competence_date=date(2026, 6, 1),
            due_date=date(2026, 6, 5),
            payment_date=date(2026, 6, 6),
            amount=amount,
            counterparty=self.counterparty,
            category=self.category,
            chart_account=self.chart_account,
            cost_center=self.cost_center,
            work=work or self.work,
            description=description,
            document_number=document_number,
            payment_method=payment_method,
            work_item_index=work_item_index,
            status=status,
            source="telegram",
            needs_review=False,
        )

    def load_export_sheet(self, batch):
        workbook = load_workbook(batch.file.path)
        return workbook[SHEET_NAME]

    def load_accounting_sheet(self, batch):
        workbook = load_workbook(batch.accounting_file.path)
        return workbook[ACCOUNTING_SHEET_NAME]
