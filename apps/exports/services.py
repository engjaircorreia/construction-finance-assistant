from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import logging
from pathlib import Path

from django.conf import settings
from django.core.files.base import ContentFile
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.core.log_safety import sanitize_log_payload, sanitize_log_text
from apps.payments.confirmation import resolve_budget_item
from apps.payments.models import Payment

from .models import ExportBatch
from .selectors import approved_payments_for_export, approved_payments_for_export_period


IMPORT_SHEET_NAME = "Planilha de Importação"
ACCOUNTING_SHEET_NAME = "Pagamentos"
IMPORT_HEADERS = [
    "Data de competência*",
    "Data de vencimento*",
    "Data de pagamento",
    "Valor*",
    "Pago a (Fornecedor)",
    "Descrição",
    "Número do Documento",
    "Categoria*",
    "Forma de Pagamento",
    "Quem Paga*",
    "Conta Bancária*",
    "Centro de Custo*",
    "Obra",
    "Índice Etapa / Item",
]
ACCOUNTING_HEADERS = [
    "Índice",
    "Data de Competência",
    "Data de Vencimento",
    "Data de Pagamento",
    "Valor da Parcela",
    "Valor em Aberto",
    "Valor Pago da Parcela",
    "Juros / Multas",
    "Descontos",
    "Valor Total Pago",
    "Fornecedor",
    "CNPJ / CPF do Fornecedor",
    "Dados Bancários do Fornecedor",
    "Descrição",
    "Número do Documento",
    "Categoria",
    "Plano de Contas",
    "Grupo",
    "Condição de Pagamento",
    "Forma de Pagamento",
    "Quem Paga",
    "Conta Bancária",
    "Centro de Custo",
    "Obra",
    "Índice Etapa / Item",
    "Etapa / Item",
    "Ordem de Compra",
    "Comentários",
]
SHEET_NAME = IMPORT_SHEET_NAME
HEADERS = IMPORT_HEADERS


logger = logging.getLogger(__name__)


class ExportValidationError(Exception):
    pass


@dataclass(frozen=True)
class PaymentExportRow:
    values: list


@transaction.atomic
def export_approved_payments(
    user=None,
    template_path: str | Path | None = None,
    accounting_template_path: str | Path | None = None,
    import_template_path: str | Path | None = None,
) -> ExportBatch:
    return export_payment_queryset(
        approved_payments_for_export(),
        user=user,
        template_path=template_path,
        accounting_template_path=accounting_template_path,
        import_template_path=import_template_path,
    )


@transaction.atomic
def export_approved_payments_for_period(
    period_start,
    period_end,
    user=None,
    template_path: str | Path | None = None,
    accounting_template_path: str | Path | None = None,
    import_template_path: str | Path | None = None,
) -> ExportBatch:
    return export_payment_queryset(
        approved_payments_for_export_period(period_start, period_end),
        user=user,
        template_path=template_path,
        accounting_template_path=accounting_template_path,
        import_template_path=import_template_path,
        notes=f"Monthly close: {period_start:%m/%Y}",
        period_start=period_start,
        period_end=period_end,
    )


@transaction.atomic
def export_payment_queryset(
    payments_queryset,
    user=None,
    template_path: str | Path | None = None,
    accounting_template_path: str | Path | None = None,
    import_template_path: str | Path | None = None,
    notes: str = "",
    period_start=None,
    period_end=None,
) -> ExportBatch:
    import_template_path = Path(import_template_path or template_path or settings.PAYMENT_IMPORT_TEMPLATE)
    accounting_template_path = Path(accounting_template_path or settings.PAYMENT_ACCOUNTING_TEMPLATE)
    payments = list(
        payments_queryset.select_related(
            "counterparty",
            "counterparty__default_category",
            "counterparty__default_chart_account",
            "category",
            "chart_account",
            "cost_center",
            "work",
        )
        .order_by("payment_date", "id")
    )
    validate_payments_can_be_exported(payments)
    batch = ExportBatch.objects.create(
        status=ExportBatch.Status.PROCESSING,
        template_path=str(import_template_path),
        import_template_path=str(import_template_path),
        accounting_template_path=str(accounting_template_path),
        generated_by=user,
        notes=notes,
        period_start=period_start,
        period_end=period_end,
    )
    try:
        generated_at = timezone.now()
        timestamp = generated_at.strftime("%Y%m%d_%H%M%S")

        import_workbook = build_import_workbook(payments, import_template_path)
        import_content = workbook_bytes(import_workbook)
        import_filename = f"payments_importacao_{timestamp}.xlsx"
        batch.import_file.save(import_filename, ContentFile(import_content), save=False)
        batch.file.save(import_filename, ContentFile(import_content), save=False)

        accounting_workbook = build_accounting_workbook(payments, accounting_template_path)
        accounting_filename = f"payments_exportacao_contador_{timestamp}.xlsx"
        batch.accounting_file.save(accounting_filename, ContentFile(workbook_bytes(accounting_workbook)), save=False)

        batch.status = ExportBatch.Status.GENERATED
        batch.records_count = len(payments)
        batch.generated_at = generated_at
        batch.save()
        batch.payments.set(payments)
        logger.info(
            "Payment export generated: %s",
            sanitize_log_payload(
                {
                    "batch_id": batch.pk,
                    "records_count": batch.records_count,
                    "period_start": period_start.isoformat() if period_start else "",
                    "period_end": period_end.isoformat() if period_end else "",
                    "generated_by_id": user.pk if user else None,
                }
            ),
        )
    except Exception as exc:
        batch.status = ExportBatch.Status.ERROR
        batch.error_message = sanitize_log_text(str(exc), limit=300)
        batch.save(update_fields=["status", "error_message", "updated_at"])
        logger.warning(
            "Payment export failed: %s",
            sanitize_log_payload(
                {
                    "batch_id": batch.pk,
                    "error_class": exc.__class__.__name__,
                    "error": str(exc),
                    "period_start": period_start.isoformat() if period_start else "",
                    "period_end": period_end.isoformat() if period_end else "",
                }
            ),
        )
        raise
    return batch


def build_payments_workbook(payments: list[Payment], template_path: str | Path):
    return build_import_workbook(payments, template_path)


def build_import_workbook(payments: list[Payment], template_path: str | Path):
    workbook = load_workbook(template_path)
    if IMPORT_SHEET_NAME not in workbook.sheetnames:
        raise ExportValidationError(f"Required sheet not found: {IMPORT_SHEET_NAME}")
    worksheet = workbook[IMPORT_SHEET_NAME]
    validate_headers(worksheet, IMPORT_HEADERS)
    clear_date_rows(worksheet, len(IMPORT_HEADERS))
    for index, payment in enumerate(payments, start=2):
        copy_row_style(worksheet, source_row=2, target_row=index, column_count=len(IMPORT_HEADERS))
        row = payment_to_import_row(payment)
        for column, value in enumerate(row.values, start=1):
            worksheet.cell(row=index, column=column).value = value
    return workbook


def build_accounting_workbook(payments: list[Payment], template_path: str | Path):
    workbook = load_workbook(template_path)
    if ACCOUNTING_SHEET_NAME not in workbook.sheetnames:
        raise ExportValidationError(f"Required sheet not found: {ACCOUNTING_SHEET_NAME}")
    worksheet = workbook[ACCOUNTING_SHEET_NAME]
    validate_headers(worksheet, ACCOUNTING_HEADERS)
    clear_date_rows(worksheet, len(ACCOUNTING_HEADERS))
    for index, payment in enumerate(payments, start=2):
        copy_row_style(worksheet, source_row=2, target_row=index, column_count=len(ACCOUNTING_HEADERS))
        row = payment_to_accounting_row(payment, row_number=index - 1)
        for column, value in enumerate(row.values, start=1):
            worksheet.cell(row=index, column=column).value = value
    return workbook


def workbook_bytes(workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def validate_payments_can_be_exported(payments: list[Payment]) -> None:
    errors = []
    for payment in payments:
        missing = payment_missing_required_fields(payment)
        if missing:
            errors.append(f"Payment {payment.pk}: {', '.join(missing)}")
    if errors:
        raise ExportValidationError("Missing required fields. " + " | ".join(errors))


def validate_headers(worksheet, expected_headers: list[str]) -> None:
    headers = [worksheet.cell(row=1, column=column).value for column in range(1, len(expected_headers) + 1)]
    if headers != expected_headers:
        raise ExportValidationError("Payment template headers do not match the expected format.")


def clear_date_rows(worksheet, column_count: int) -> None:
    for row in range(2, worksheet.max_row + 1):
        for column in range(1, column_count + 1):
            worksheet.cell(row=row, column=column).value = None


def copy_row_style(worksheet, source_row: int, target_row: int, column_count: int) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column in range(1, column_count + 1):
        source = worksheet.cell(row=source_row, column=column)
        target = worksheet.cell(row=target_row, column=column)
        target._style = copy(source._style)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def payment_to_export_row(payment: Payment) -> PaymentExportRow:
    return payment_to_import_row(payment)


def payment_to_import_row(payment: Payment) -> PaymentExportRow:
    date = payment_export_date(payment)
    return PaymentExportRow(
        values=[
            date["competence_date"],
            date["due_date"],
            date["payment_date"],
            format_amount(payment.amount),
            date["counterparty_name"],
            date["description"],
            payment.document_number,
            date["category"].name,
            payment.payment_method,
            date["payer"],
            date["bank_account"],
            date["cost_center"].name,
            date["work_name"],
            payment.work_item_index,
        ]
    )


def payment_to_accounting_row(payment: Payment, row_number: int) -> PaymentExportRow:
    date = payment_export_date(payment)
    amount = format_amount(payment.amount)
    budget_item = resolve_budget_item(payment)
    return PaymentExportRow(
        values=[
            str(row_number),
            date["competence_date"],
            date["due_date"],
            date["payment_date"],
            amount,
            Decimal("0.00"),
            amount,
            Decimal("0.00"),
            Decimal("0.00"),
            amount,
            date["counterparty_name"],
            date["counterparty_document"],
            "",
            date["description"],
            payment.document_number,
            date["category"].name,
            date["chart_account_name"],
            date["group_name"],
            "À Vista",
            payment.payment_method,
            date["payer"],
            date["bank_account"],
            date["cost_center"].name,
            date["work_name"] or "-",
            payment.work_item_index or "-",
            budget_item.description if budget_item else "-",
            "",
            "",
        ]
    )


def payment_export_date(payment: Payment) -> dict:
    date = payment_export_base_date(payment)
    missing = payment_missing_required_fields(payment)
    if missing:
        raise ExportValidationError(f"Payment {payment.pk} is missing required fields: {', '.join(missing)}")
    return {
        **date,
        "description": (payment.description or "")[:200],
        "group_name": group_name_for_category(date["category"].name),
    }


def payment_export_base_date(payment: Payment) -> dict:
    category = payment.category or getattr(payment.counterparty, "default_category", None)
    cost_center = payment.cost_center
    chart_account = payment.chart_account or getattr(payment.counterparty, "default_chart_account", None)
    competence_date = payment.competence_date or payment.payment_date or payment.due_date
    due_date = payment.due_date or payment.payment_date or payment.competence_date
    payer = payment.payer or settings.DEFAULT_PAYER
    bank_account = payment.bank_account or settings.DEFAULT_BANK_ACCOUNT

    counterparty_name = payment.counterparty.name if payment.counterparty else ""
    counterparty_document = payment.counterparty.primary_document if payment.counterparty else ""
    work_name = payment.work.name if payment.work else ""
    chart_account_name = chart_account.name if chart_account else category.name if category else ""
    return {
        "category": category,
        "cost_center": cost_center,
        "competence_date": competence_date,
        "due_date": due_date,
        "payment_date": payment.payment_date,
        "payer": payer,
        "bank_account": bank_account,
        "counterparty_name": counterparty_name,
        "counterparty_document": counterparty_document,
        "work_name": work_name,
        "chart_account_name": chart_account_name,
    }


def payment_missing_required_fields(payment: Payment) -> list[str]:
    date = payment_export_base_date(payment)
    missing = []
    if not date["competence_date"]:
        missing.append("Accrual date")
    if not date["due_date"]:
        missing.append("Due date")
    if payment.amount is None:
        missing.append("Amount")
    if not date["category"]:
        missing.append("Category")
    if not date["payer"]:
        missing.append("Payer")
    if not date["bank_account"]:
        missing.append("Bank account")
    if not date["cost_center"]:
        missing.append("Cost center")
    return missing


def group_name_for_category(category_name: str) -> str:
    normalized = str(category_name or "").casefold()
    if "mão de project" in normalized or "mao de project" in normalized:
        return "Labor (Terceirizada)"
    if normalized in {
        "aluguel",
        "consultoria",
        "distribuição de lucros",
        "distribuicao de lucros",
        "outras despesas",
        "payment de empréstimo",
        "payment de emprestimo",
        "payment de financiamento",
        "transporte",
    }:
        return "Other"
    return category_name


def format_amount(amount) -> Decimal:
    return Decimal(amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
