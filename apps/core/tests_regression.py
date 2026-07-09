from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

from django.contrib.auth import get_user_model
from django.db import IntegrityError, transaction
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.accounts.models import AuthorizedTelegramUser
from apps.banking.models import OfxFile, OfxTransaction
from apps.banking.reconciliation import reconcile_ofx_transaction
from apps.counterparties.importers import import_counterparty_workbooks
from apps.counterparties.models import BudgetItem, Category, Counterparty, CostCenter, Work
from apps.documents.models import UploadedFile
from apps.exports.models import ExportBatch
from apps.exports.services import ACCOUNTING_HEADERS, ACCOUNTING_SHEET_NAME, HEADERS, SHEET_NAME, export_approved_payments
from apps.payments.confirmation import PaymentConfirmationError, approve_payment
from apps.payments.models import Payment, PaymentConfirmation
from apps.telegrambot.models import TelegramDraft
from apps.telegrambot.services import TelegramAttachment, TelegramIntakeService, TelegramSender


class MainRegressionTests(TestCase):
    def setUp(self):
        self.tempdir = TemporaryDirectory()
        self.override = override_settings(
            MEDIA_ROOT=Path(self.tempdir.name),
            TELEGRAM_ALLOWED_USER_IDS=[],
            DEFAULT_PAYER="In Plant Engenharia",
            DEFAULT_BANK_ACCOUNT="Banco Principal",
        )
        self.override.enable()

    def tearDown(self):
        self.override.disable()
        self.tempdir.cleanup()

    def test_unauthorized_telegram_user_is_blocked_without_side_effects(self):
        result = TelegramIntakeService().process_text(
            sender=TelegramSender(telegram_user_id=999999, name="No autorizado"),
            message_id=1,
            text="paguei 100 de cimento",
        )

        self.assertFalse(result.authorized)
        self.assertEqual(result.reply_text, "Unauthorized access.")
        self.assertEqual(UploadedFile.objects.count(), 0)
        self.assertEqual(Payment.objects.count(), 0)

    def test_payment_is_not_exported_without_approval(self):
        Payment.objects.create(
            amount=Decimal("100.00"),
            status=Payment.Status.RECEIVED,
            source="telegram",
            needs_review=True,
        )

        batch = export_approved_payments()
        worksheet = load_workbook(batch.file.path)[SHEET_NAME]

        self.assertEqual(batch.records_count, 0)
        self.assertEqual(batch.payments.count(), 0)
        self.assertIsNone(worksheet.cell(row=2, column=4).value)

    def test_new_counterparty_requires_registration_before_payment_approval(self):
        payment = Payment.objects.create(
            amount=Decimal("100.00"),
            status=Payment.Status.RECEIVED,
            source="telegram",
            needs_review=True,
            raw_payload={"counterparty_candidate": {"name": "Vendor Novo", "document": "12345678000199"}},
        )

        with self.assertRaises(PaymentConfirmationError):
            approve_payment(payment.pk, telegram_user_id=123, message="aprovar")
        payment.refresh_from_db()

        self.assertEqual(payment.status, Payment.Status.PENDING_REGISTRATION)
        self.assertEqual(Counterparty.objects.count(), 0)
        self.assertEqual(PaymentConfirmation.objects.count(), 0)

    def test_supplier_import_avoids_duplicate_counterparties_by_document(self):
        workbook_path = self.write_counterparty_workbook(
            [
                ["ACME Materiais", None, "12345678000199", "Active", "Legal entity"],
                ["ACME Comércio", None, "12.345.678/0001-99", "Active", "Legal entity"],
            ]
        )

        report = import_counterparty_workbooks(supplier_path=workbook_path)

        self.assertEqual(report.created, 1)
        self.assertEqual(report.updated, 1)
        self.assertEqual(Counterparty.objects.count(), 1)
        self.assertTrue(Counterparty.objects.get().aliases.filter(normalized_name="acme comercio").exists())

    def test_ofx_import_storage_avoids_duplicate_transactions_by_fitid(self):
        ofx_file = OfxFile.objects.create(original_filename="extrato.ofx")
        OfxTransaction.objects.create(
            ofx_file=ofx_file,
            fitid="FIT-UNICO",
            posted_at=date(2026, 6, 24),
            amount=Decimal("-100.00"),
        )

        with self.assertRaises(IntegrityError), transaction.atomic():
            OfxTransaction.objects.create(
                ofx_file=ofx_file,
                fitid="FIT-UNICO",
                posted_at=date(2026, 6, 24),
                amount=Decimal("-100.00"),
            )

    def test_export_generates_spreadsheet_with_expected_columns(self):
        category = Category.objects.create(name="Materiais", normalized_name="materiais")
        cost_center = CostCenter.objects.create(name="Project", normalized_name="project")
        counterparty = Counterparty.objects.create(name="ACME Materiais", normalized_name="acme materiais")
        Payment.objects.create(
            competence_date=date(2026, 6, 1),
            due_date=date(2026, 6, 5),
            payment_date=date(2026, 6, 6),
            amount=Decimal("150.50"),
            counterparty=counterparty,
            category=category,
            cost_center=cost_center,
            description="Compra de materiais",
            status=Payment.Status.APPROVED,
            source="telegram",
            needs_review=False,
        )

        batch = export_approved_payments()
        worksheet = load_workbook(batch.file.path)[SHEET_NAME]
        accounting_worksheet = load_workbook(batch.accounting_file.path)[ACCOUNTING_SHEET_NAME]
        headers = [worksheet.cell(row=1, column=column).value for column in range(1, len(HEADERS) + 1)]
        accounting_headers = [
            accounting_worksheet.cell(row=1, column=column).value
            for column in range(1, len(ACCOUNTING_HEADERS) + 1)
        ]

        self.assertEqual(batch.status, ExportBatch.Status.GENERATED)
        self.assertEqual(headers, HEADERS)
        self.assertEqual(accounting_headers, ACCOUNTING_HEADERS)
        self.assertEqual(worksheet.cell(row=2, column=4).value, 150.5)
        self.assertEqual(accounting_worksheet.cell(row=2, column=10).value, 150.5)

    def test_ofx_credits_are_ignored_and_do_not_create_payments(self):
        ofx_file = OfxFile.objects.create(original_filename="extrato.ofx")
        transaction_record = OfxTransaction.objects.create(
            ofx_file=ofx_file,
            fitid="CREDITO-1",
            posted_at=date(2026, 6, 24),
            amount=Decimal("500.00"),
            memo="PIX RECEBIDO CLIENTE",
        )

        result = reconcile_ofx_transaction(transaction_record)
        transaction_record.refresh_from_db()

        self.assertEqual(result["classification"], "ignored_credits")
        self.assertEqual(transaction_record.status, OfxTransaction.Status.IGNORED)
        self.assertEqual(Payment.objects.count(), 0)

    def write_counterparty_workbook(self, rows):
        path = Path(self.tempdir.name) / "vendors.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Vendors"
        sheet.append(["Name", "CPF", "CNPJ", "Status", "Type"])
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        return path


class CompleteFlowValidationTests(TestCase):
    def setUp(self):
        self.tempdir = TemporaryDirectory()
        self.override = override_settings(
            MEDIA_ROOT=Path(self.tempdir.name),
            TELEGRAM_ALLOWED_USER_IDS=[],
            OPENAI_API_KEY="",
            OPENAI_DRAFT_REARRANGE_ENABLED=False,
        )
        self.override.enable()
        self.service = TelegramIntakeService()
        self.sender = TelegramSender(telegram_user_id=123, name="Jair", username="JairCorreia")
        AuthorizedTelegramUser.objects.create(telegram_user_id=123, name="Jair", username="JairCorreia")
        self.company_cost_center = CostCenter.objects.create(name="Company", normalized_name="empresa")
        self.work_cost_center = CostCenter.objects.create(name="Project", normalized_name="project")

    def tearDown(self):
        self.override.disable()
        self.tempdir.cleanup()

    def test_known_pdf_work_and_counterparty_can_be_finalized_then_approved(self):
        category = Category.objects.create(name="Materiais", normalized_name="materiais")
        work = Work.objects.create(name="Sertãozinho", normalized_name="sertaozinho")
        BudgetItem.objects.create(
            work=work,
            index="1.4",
            item_type=BudgetItem.ItemType.SUBSTAGE,
            description="CALÇADA",
            normalized_description="calcada",
        )
        counterparty = Counterparty.objects.create(
            name="ACME Materiais",
            normalized_name="acme materiais",
            default_category=category,
        )

        self.send_pdf(
            "Receipt de Payment Pix Amount: R$ 100,00 Realized em: 24/06/2026 "
            "Name do destinatário: ACME Materiais project Sertãozinho"
        )
        draft = TelegramDraft.objects.get()

        self.assertEqual(draft.counterparty, counterparty)
        self.assertEqual(draft.work, work)
        self.assertEqual(draft.cost_center, self.work_cost_center)
        self.assertEqual(Payment.objects.count(), 0)

        result = self.service.finalize_draft(draft.pk, self.sender)
        payment = result.payment

        self.assertIsNotNone(payment)
        self.assertEqual(payment.status, Payment.Status.PENDING_CONFIRMATION)
        self.assertTrue(payment.needs_review)
        self.assertEqual(payment.counterparty, counterparty)
        self.assertEqual(payment.work, work)

        approve_payment(payment.pk, telegram_user_id=123, message="aprovar")
        payment.refresh_from_db()

        self.assertEqual(payment.status, Payment.Status.APPROVED)
        self.assertFalse(payment.needs_review)

    def test_new_counterparty_from_pdf_is_registered_by_telegram_before_approval(self):
        result = self.send_pdf(
            "Receipt de Payment Pix Amount: R$ 250,00 Realized em: 24/06/2026 "
            "Name do destinatário: Novo Vendor Ltda"
        )
        draft = TelegramDraft.objects.get()

        self.assertIn("Suggested registration", result.reply_text)
        self.assertIsNone(draft.counterparty)
        self.assertIn("counterparty_candidate", draft.raw_payload)

        register_result = self.service.register_counterparty_from_draft(
            draft.pk,
            Counterparty.Kind.SUPPLIER,
            self.sender,
        )
        draft.refresh_from_db()
        counterparty = Counterparty.objects.get(normalized_name="novo vendor ltda")

        self.assertIn("Vendor registered", register_result.reply_text)
        self.assertEqual(draft.counterparty, counterparty)
        self.assertEqual(Payment.objects.count(), 0)

        finalize_result = self.service.finalize_draft(draft.pk, self.sender)
        payment = finalize_result.payment

        self.assertEqual(payment.status, Payment.Status.PENDING_CONFIRMATION)
        approve_payment(payment.pk, telegram_user_id=123, message="aprovar")
        payment.refresh_from_db()
        self.assertEqual(payment.status, Payment.Status.APPROVED)

    def test_new_work_from_text_is_registered_then_finalized_with_empty_budget_index(self):
        counterparty = Counterparty.objects.create(name="Vendor Project", normalized_name="vendor project")
        self.send_pdf(
            "Receipt de Payment Pix Amount: R$ 500,00 Realized em: 24/06/2026 "
            "Name do destinatário: Vendor Project"
        )
        work_message = self.service.process_text(self.sender, 200, "project Name Novo")
        draft = TelegramDraft.objects.get()

        self.assertEqual(draft.counterparty, counterparty)
        self.assertIn("Suggested project not registered: Name Novo", work_message.reply_text)
        self.assertEqual(draft.cost_center, self.company_cost_center)

        self.service.register_work_from_draft(draft.pk, self.sender)
        draft.refresh_from_db()

        self.assertEqual(draft.work.name, "Name Novo")
        self.assertEqual(draft.cost_center, self.work_cost_center)

        result = self.service.finalize_draft(draft.pk, self.sender)
        payment = result.payment

        self.assertEqual(payment.status, Payment.Status.PENDING_CONFIRMATION)
        self.assertEqual(payment.work.name, "Name Novo")
        self.assertEqual(payment.work_item_index, "")
        approve_payment(payment.pk, telegram_user_id=123, message="aprovar")
        payment.refresh_from_db()
        self.assertEqual(payment.status, Payment.Status.APPROVED)

    def test_tax_expense_without_work_uses_company_cost_center_and_can_be_approved(self):
        counterparty = Counterparty.objects.create(name="Income Federal", normalized_name="income federal")

        self.send_pdf(
            "Receipt de Payment Pix Amount: R$ 180,00 Realized em: 24/06/2026 "
            "Name do destinatário: Income Federal imposto simples nacional"
        )
        draft = TelegramDraft.objects.get()

        self.assertEqual(draft.counterparty, counterparty)
        self.assertEqual(draft.cost_center, self.company_cost_center)
        self.assertIsNone(draft.work)

        result = self.service.finalize_draft(draft.pk, self.sender)
        payment = result.payment

        self.assertEqual(payment.status, Payment.Status.PENDING_CONFIRMATION)
        self.assertEqual(payment.cost_center, self.company_cost_center)
        self.assertIsNone(payment.work)
        approve_payment(payment.pk, telegram_user_id=123, message="aprovar")
        payment.refresh_from_db()
        self.assertEqual(payment.status, Payment.Status.APPROVED)

    def test_web_draft_edit_finalize_creates_payment_visible_in_payments(self):
        User = get_user_model()
        user = User.objects.create_user(username="jair", password="senha-forte")
        self.client.force_login(user)
        category = Category.objects.create(name="Services", normalized_name="servicos")
        counterparty = Counterparty.objects.create(name="Vendor Web", normalized_name="vendor web")
        work = Work.objects.create(name="Tacima", normalized_name="tacima")
        draft = TelegramDraft.objects.create(telegram_user_id=123, sender_name="Jair")

        list_response = self.client.get(reverse("internal_telegram_drafts"))
        self.assertContains(list_response, f"#{draft.pk}")

        update_response = self.client.post(
            reverse("internal_telegram_draft_update", args=[draft.pk]),
            {
                "payment_date": "2026-06-24",
                "amount": "300.00",
                "counterparty": str(counterparty.pk),
                "description": "Payment editado pela web",
                "category": str(category.pk),
                "payment_method": "PIX",
                "cost_center": str(self.work_cost_center.pk),
                "work": str(work.pk),
                "work_item_index": "",
            },
        )
        self.assertEqual(update_response.status_code, 302)

        finalize_response = self.client.post(reverse("internal_telegram_draft_finalize", args=[draft.pk]))
        payment = Payment.objects.get()

        self.assertEqual(finalize_response.status_code, 302)
        self.assertEqual(finalize_response["Location"], reverse("internal_payment_detail", args=[payment.pk]))
        self.assertEqual(payment.status, Payment.Status.PENDING_CONFIRMATION)
        self.assertEqual(payment.description, "Payment editado pela web")

        payments_response = self.client.get(
            reverse("internal_pending_payments"),
            {"status": "all", "date_inicio": "2026-06-01", "date_fim": "2026-06-30"},
        )
        self.assertContains(payments_response, "Vendor Web")
        self.assertContains(payments_response, "R$ 300,00")

    def test_security_flow_blocks_anonymous_logout_works_old_callbacks_do_not_auto_approve(self):
        response = self.client.get(reverse("internal_telegram_drafts"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

        User = get_user_model()
        user = User.objects.create_user(username="socio", password="senha-forte")
        self.client.force_login(user)
        logout_response = self.client.post(reverse("logout"))
        self.assertEqual(logout_response.status_code, 302)
        self.assertEqual(logout_response["Location"], "/accounts/login/")

        payment = Payment.objects.create(source="telegram", status=Payment.Status.PENDING_CONFIRMATION)
        finalized = TelegramDraft.objects.create(
            telegram_user_id=123,
            status=TelegramDraft.Status.FINALIZED,
            finalized_payment=payment,
            raw_payload={"counterparty_candidate": {"name": "Pessoa Finalizada", "document": "", "source": "texto"}},
        )
        canceled = TelegramDraft.objects.create(
            telegram_user_id=123,
            status=TelegramDraft.Status.CANCELED,
            raw_payload={"work_candidate": {"name": "Project Cancelada", "source": "telegram"}},
        )

        counterparty_result = self.service.register_counterparty_from_draft(
            finalized.pk,
            Counterparty.Kind.SUPPLIER,
            self.sender,
        )
        work_result = self.service.register_work_from_draft(canceled.pk, self.sender)
        finalized.refresh_from_db()
        canceled.refresh_from_db()
        payment.refresh_from_db()

        self.assertIn("not active", counterparty_result.reply_text)
        self.assertIn("not active", work_result.reply_text)
        self.assertEqual(Counterparty.objects.count(), 0)
        self.assertEqual(Work.objects.count(), 0)
        self.assertEqual(payment.status, Payment.Status.PENDING_CONFIRMATION)

    def send_pdf(self, text: str, message_id: int = 100):
        return self.service.process_attachment(
            sender=self.sender,
            message_id=message_id,
            attachment=TelegramAttachment(
                file_id=f"pdf-{message_id}",
                filename=f"receipt-{message_id}.pdf",
                content_type="application/pdf",
                content=minimal_pdf_with_text(text),
                kind=UploadedFile.Kind.PDF,
            ),
        )


def minimal_pdf_with_text(text):
    escaped = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    stream = f"BT /F1 12 Tf 72 720 Td ({escaped}) Tj ET".encode("latin-1")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    chunks = [b"%PDF-1.4\n"]
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(sum(len(chunk) for chunk in chunks))
        chunks.append(f"{index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n")
    xref_offset = sum(len(chunk) for chunk in chunks)
    chunks.append(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    chunks.append(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        chunks.append(f"{offset:010d} 00000 n \n".encode("ascii"))
    chunks.append(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    return b"".join(chunks)
