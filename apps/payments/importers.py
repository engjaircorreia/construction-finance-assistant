from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import hashlib
import json

from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook

from apps.counterparties.importers import (
    clean_name,
    digits_only,
    document_type,
    normalize_header,
    normalize_text,
)
from apps.counterparties.matching import find_best_counterparty_by_normalized_name
from apps.counterparties.models import (
    Category,
    ChartOfAccount,
    CostCenter,
    Counterparty,
    CounterpartyAlias,
    CounterpartyDocument,
    Origin,
    Work,
)

from .models import Payment


HEADER_ALIASES = {
    "indice": "index",
    "date de competencia": "competence_date",
    "data de competencia": "competence_date",
    "due date": "due_date",
    "date de vencimento": "due_date",
    "data de vencimento": "due_date",
    "payment date": "payment_date",
    "date de payment": "payment_date",
    "data de pagamento": "payment_date",
    "amount da parcela": "installment_amount",
    "valor da parcela": "installment_amount",
    "amount em aberto": "open_amount",
    "valor em aberto": "open_amount",
    "amount pago da parcela": "paid_installment_amount",
    "valor pago da parcela": "paid_installment_amount",
    "amount total pago": "paid_amount",
    "valor total pago": "paid_amount",
    "vendor": "counterparty_name",
    "fornecedor": "counterparty_name",
    "favorecido": "counterparty_name",
    "payee": "counterparty_name",
    "cnpj / cpf do vendor": "counterparty_document",
    "cnpj/cpf do vendor": "counterparty_document",
    "cnpj / cpf do fornecedor": "counterparty_document",
    "cnpj/cpf do fornecedor": "counterparty_document",
    "descricao": "description",
    "description": "description",
    "numero do documento": "document_number",
    "document number": "document_number",
    "category": "category",
    "categoria": "category",
    "chart of accounts": "chart_account",
    "plano de contas": "chart_account",
    "cost center": "cost_center",
    "centro de custo": "cost_center",
    "project": "work",
    "obra": "work",
    "indice etapa / item": "work_item_index",
    "indice step / item": "work_item_index",
    "payment method": "payment_method",
    "condicao de payment": "payment_condition",
    "condicao de pagamento": "payment_condition",
    "payer": "payer",
    "conta bancaria": "bank_account",
    "bank account": "bank_account",
}

LABOR_CATEGORY_NAMES = {"mao de obra terceirizada", "mao de project terceirizada", "labor terceirizada"}
CANONICAL_LABOR_CATEGORY = "Mao de Obra Terceirizada"


@dataclass
class PaymentHistoryConflict:
    source_file: str
    row_number: int
    counterparty_name: str
    reason: str
    detail: str


@dataclass
class PaymentHistoryReport:
    files: list[str] = field(default_factory=list)
    rows_read: int = 0
    rows_skipped: int = 0
    unpaid_or_non_expense_skipped: int = 0
    payments_created: int = 0
    payments_unchanged: int = 0
    counterparties_created: int = 0
    counterparties_updated: int = 0
    categories_created: int = 0
    chart_accounts_created: int = 0
    cost_centers_created: int = 0
    works_created: int = 0
    documents_created: int = 0
    aliases_created: int = 0
    classification_rules_updated: int = 0
    conflicts: list[PaymentHistoryConflict] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "files": self.files,
            "rows_read": self.rows_read,
            "rows_skipped": self.rows_skipped,
            "unpaid_or_non_expense_skipped": self.unpaid_or_non_expense_skipped,
            "payments_created": self.payments_created,
            "payments_unchanged": self.payments_unchanged,
            "counterparties_created": self.counterparties_created,
            "counterparties_updated": self.counterparties_updated,
            "categories_created": self.categories_created,
            "chart_accounts_created": self.chart_accounts_created,
            "cost_centers_created": self.cost_centers_created,
            "works_created": self.works_created,
            "documents_created": self.documents_created,
            "aliases_created": self.aliases_created,
            "classification_rules_updated": self.classification_rules_updated,
            "conflicts": [conflict.__dict__ for conflict in self.conflicts],
        }


class PaymentHistoryImporter:
    def __init__(self, create_payments: bool = True):
        self.create_payments = create_payments
        self.report = PaymentHistoryReport()
        self.stats = defaultdict(
            lambda: {
                "kind": Counter(),
                "category": Counter(),
                "chart_account": Counter(),
                "cost_center": Counter(),
                "work": Counter(),
            }
        )

    def import_path(self, path: str | Path, dry_run: bool = False) -> PaymentHistoryReport:
        path = Path(path)
        with transaction.atomic():
            paths = sorted(path.glob("*.xlsx")) if path.is_dir() else [path]
            for workbook_path in paths:
                self.import_workbook(workbook_path)
            self.apply_classification_rules()
            if dry_run:
                transaction.set_rollback(True)
        return self.report

    def import_workbook(self, path: Path) -> None:
        self.report.files.append(str(path))
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            sheet.reset_dimensions()
            header_row_number, header = find_header(sheet)
            if not header:
                continue
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
                start=header_row_number + 1,
            ):
                raw = row_to_dict(header, row)
                if not any(raw.values()):
                    continue
                self.report.rows_read += 1
                self.import_row(path, row_number, raw)

    def import_row(self, path: Path, row_number: int, raw: dict) -> None:
        name = clean_name(raw.get("counterparty_name"))
        normalized_name = normalize_text(name)
        if not name or normalized_name == "-":
            self.report.rows_skipped += 1
            return

        category_name = canonical_lookup_name(raw.get("category"))
        chart_account_name = canonical_lookup_name(raw.get("chart_account"))
        category = self.get_lookup(Category, category_name, "categories_created")
        chart_account = self.get_lookup(ChartOfAccount, chart_account_name, "chart_accounts_created")
        cost_center = self.get_lookup(CostCenter, raw.get("cost_center"), "cost_centers_created")
        work = self.get_work(raw.get("work"))

        document = digits_only(raw.get("counterparty_document"))
        inferred_kind = infer_kind(raw.get("category"), document)
        counterparty = self.get_or_create_counterparty(
            name=name,
            normalized_name=normalized_name,
            document=document,
            inferred_kind=inferred_kind,
        )

        self.ensure_document(counterparty, document)
        self.ensure_alias(counterparty, name, normalized_name)
        self.collect_classification(counterparty, inferred_kind, category, chart_account, cost_center, work)

        amount = parse_decimal(raw.get("paid_amount"))
        payment_date = parse_date(raw.get("payment_date"))
        if amount <= 0 or payment_date is None:
            self.report.unpaid_or_non_expense_skipped += 1
            return
        if not self.create_payments:
            return

        history_key = build_history_key(path, raw.get("index") or row_number)
        if Payment.objects.filter(source=Origin.HISTORICAL, raw_payload__history_key=history_key).exists():
            self.report.payments_unchanged += 1
            return

        Payment.objects.create(
            competence_date=parse_date(raw.get("competence_date")),
            due_date=parse_date(raw.get("due_date")),
            payment_date=payment_date,
            amount=amount,
            counterparty=counterparty,
            description=clean_name(raw.get("description"))[:255],
            document_number=clean_name(raw.get("document_number")),
            category=category,
            chart_account=chart_account,
            payment_method=clean_name(raw.get("payment_method")),
            payer=clean_name(raw.get("payer")) or settings.DEFAULT_PAYER,
            bank_account=clean_name(raw.get("bank_account")) or settings.DEFAULT_BANK_ACCOUNT,
            cost_center=cost_center,
            work=work,
            work_item_index=clean_name(raw.get("work_item_index")),
            source=Origin.HISTORICAL,
            status=Payment.Status.POSTED,
            confidence=1,
            needs_review=False,
            raw_payload={
                "history_key": history_key,
                "source_file": path.name,
                "source_row": row_number,
                "raw": json_safe(raw),
            },
        )
        self.report.payments_created += 1

    def get_lookup(self, model, value: object, report_field: str):
        name = clean_lookup_name(value)
        if not name:
            return None
        normalized_name = normalize_text(name)
        obj, created = model.objects.get_or_create(
            normalized_name=normalized_name,
            defaults={"name": name},
        )
        if created:
            setattr(self.report, report_field, getattr(self.report, report_field) + 1)
        return obj

    def get_work(self, value: object) -> Work | None:
        name = clean_lookup_name(value)
        if not name:
            return None
        normalized_name = normalize_text(name)
        obj, created = Work.objects.get_or_create(
            normalized_name=normalized_name,
            defaults={"name": name},
        )
        if created:
            self.report.works_created += 1
        return obj

    def get_or_create_counterparty(
        self,
        name: str,
        normalized_name: str,
        document: str,
        inferred_kind: str,
    ) -> Counterparty:
        counterparty = find_counterparty(document, normalized_name, inferred_kind)
        if counterparty is None:
            counterparty = Counterparty.objects.create(
                name=name,
                normalized_name=normalized_name,
                kind=inferred_kind,
                person_type=person_type_from_document(document),
                primary_document=document,
                source=Origin.HISTORICAL,
                confidence=1 if document else 0,
            )
            self.report.counterparties_created += 1
            return counterparty

        changed = False
        if not counterparty.primary_document and document:
            counterparty.primary_document = document
            changed = True
        if counterparty.person_type == Counterparty.PersonType.UNKNOWN:
            person_type = person_type_from_document(document)
            if person_type != Counterparty.PersonType.UNKNOWN:
                counterparty.person_type = person_type
                changed = True
        if counterparty.source == Origin.MANUAL:
            counterparty.source = Origin.HISTORICAL
            changed = True
        if changed:
            counterparty.save()
            self.report.counterparties_updated += 1
        return counterparty

    def ensure_document(self, counterparty: Counterparty, document: str) -> None:
        if not document or CounterpartyDocument.objects.filter(number=document).exists():
            return
        is_primary = counterparty.primary_document == document and not counterparty.documents.filter(is_primary=True).exists()
        CounterpartyDocument.objects.create(
            counterparty=counterparty,
            document_type=document_type(document),
            number=document,
            source=Origin.HISTORICAL,
            confidence=1,
            is_primary=is_primary,
        )
        self.report.documents_created += 1

    def ensure_alias(self, counterparty: Counterparty, name: str, normalized_name: str) -> None:
        if not normalized_name or normalized_name == counterparty.normalized_name:
            return
        _, created = CounterpartyAlias.objects.get_or_create(
            counterparty=counterparty,
            normalized_name=normalized_name,
            defaults={"name": name, "source": Origin.HISTORICAL},
        )
        if created:
            self.report.aliases_created += 1

    def collect_classification(self, counterparty, inferred_kind, category, chart_account, cost_center, work) -> None:
        stats = self.stats[counterparty.pk]
        if inferred_kind:
            stats["kind"][inferred_kind] += 1
        if category:
            stats["category"][category.pk] += 1
        if chart_account:
            stats["chart_account"][chart_account.pk] += 1
        if cost_center:
            stats["cost_center"][cost_center.pk] += 1
        if work:
            stats["work"][work.pk] += 1

    def apply_classification_rules(self) -> None:
        for counterparty_id, stats in self.stats.items():
            counterparty = Counterparty.objects.get(pk=counterparty_id)
            changed = False
            kind = most_common_key(stats["kind"])
            category_id = most_common_key(stats["category"])
            chart_account_id = most_common_key(stats["chart_account"])
            cost_center_id = most_common_key(stats["cost_center"])
            work_id = most_common_key(stats["work"])
            if kind and counterparty.kind != kind:
                old_kind = counterparty.kind
                counterparty.kind = kind
                changed = True
                self.report.conflicts.append(
                    PaymentHistoryConflict(
                        source_file="histórico agregado",
                        row_number=0,
                        counterparty_name=counterparty.name,
                        reason="kind_changed_by_history",
                        detail=f"existing={old_kind}; inferred={kind}",
                    )
                )
            if category_id and counterparty.default_category_id != category_id:
                counterparty.default_category_id = category_id
                changed = True
            if chart_account_id and counterparty.default_chart_account_id != chart_account_id:
                counterparty.default_chart_account_id = chart_account_id
                changed = True
            if cost_center_id and counterparty.default_cost_center_id != cost_center_id:
                counterparty.default_cost_center_id = cost_center_id
                changed = True
            if work_id and counterparty.default_work_id != work_id:
                counterparty.default_work_id = work_id
                changed = True
            if changed:
                counterparty.save()
                self.report.classification_rules_updated += 1


def import_payment_history(path: str | Path | None = None, dry_run: bool = False, create_payments: bool = True):
    importer = PaymentHistoryImporter(create_payments=create_payments)
    return importer.import_path(path or settings.PAYMENTS_HISTORY_DIR, dry_run=dry_run)


def find_header(sheet) -> tuple[int, dict[int, str]]:
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        header = {}
        for index, value in enumerate(row):
            field = HEADER_ALIASES.get(normalize_header(value))
            if field:
                header[index] = field
        if "counterparty_name" in header.values():
            return row_number, header
    return 0, {}


def row_to_dict(header: dict[int, str], row: tuple) -> dict:
    date = {}
    for index, field in header.items():
        value = row[index] if index < len(row) else None
        if value is not None and str(value).strip() != "":
            date[field] = value
    return date


def clean_lookup_name(value: object) -> str:
    name = clean_name(value)
    if not name or normalize_text(name) == "-":
        return ""
    return name


def parse_decimal(value: object) -> Decimal:
    if value is None or str(value).strip() == "":
        return Decimal("0.00")
    try:
        amount = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def parse_date(value: object) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def infer_kind(category: object, document: str) -> str:
    if normalize_text(category) in LABOR_CATEGORY_NAMES:
        return Counterparty.Kind.WORKER
    return Counterparty.Kind.SUPPLIER


def canonical_lookup_name(value: object) -> str:
    name = clean_lookup_name(value)
    if normalize_text(name) in LABOR_CATEGORY_NAMES:
        return CANONICAL_LABOR_CATEGORY
    return name


def person_type_from_document(document: str) -> str:
    if len(document) == 11:
        return Counterparty.PersonType.INDIVIDUAL
    if len(document) == 14:
        return Counterparty.PersonType.COMPANY
    return Counterparty.PersonType.UNKNOWN


def find_counterparty(document: str, normalized_name: str, kind: str = "") -> Counterparty | None:
    if document:
        document_record = (
            CounterpartyDocument.objects.select_related("counterparty").filter(number=document).first()
        )
        if document_record:
            return document_record.counterparty
        counterparty = Counterparty.objects.filter(primary_document=document).first()
        if counterparty:
            return counterparty
    return find_best_counterparty_by_normalized_name(normalized_name, kind=kind)


def most_common_key(counter: Counter):
    if not counter:
        return None
    return counter.most_common(1)[0][0]


def build_history_key(path: Path, row_index: object) -> str:
    raw = f"{path.name}:{row_index}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def json_safe(raw: dict) -> dict:
    safe = {}
    for key, value in raw.items():
        if isinstance(value, (date, datetime)):
            safe[key] = value.isoformat()
        else:
            safe[key] = str(value)
    return safe
