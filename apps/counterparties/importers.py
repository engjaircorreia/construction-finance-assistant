from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
import unicodedata

from django.db import transaction
from openpyxl import load_workbook

from .models import BudgetItem, Counterparty, CounterpartyAlias, CounterpartyDocument, Origin, Work


HEADER_ALIASES = {
    "name": "name",
    "name*": "name",
    "razao social": "legal_name",
    "cpf": "cpf",
    "cnpj": "cnpj",
    "cpf/cnpj": "document",
    "type": "person_type",
    "type*": "person_type",
    "status": "status",
    "email": "email",
    "e-mail": "email",
    "telefones": "phone",
    "telefone": "phone",
    "conta bancaria": "bank_account",
    "cargo": "role",
    "fonte": "source_note",
    "canal": "source_note",
}


@dataclass
class ImportConflict:
    row_number: int
    name: str
    document: str
    reason: str
    detail: str


@dataclass
class ImportReport:
    files: list[str] = field(default_factory=list)
    rows_read: int = 0
    rows_skipped: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    documents_created: int = 0
    aliases_created: int = 0
    conflicts: list[ImportConflict] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "files": self.files,
            "rows_read": self.rows_read,
            "rows_skipped": self.rows_skipped,
            "created": self.created,
            "updated": self.updated,
            "unchanged": self.unchanged,
            "documents_created": self.documents_created,
            "aliases_created": self.aliases_created,
            "conflicts": [conflict.__dict__ for conflict in self.conflicts],
        }

    def add(self, other: "ImportReport") -> None:
        self.files.extend(other.files)
        self.rows_read += other.rows_read
        self.rows_skipped += other.rows_skipped
        self.created += other.created
        self.updated += other.updated
        self.unchanged += other.unchanged
        self.documents_created += other.documents_created
        self.aliases_created += other.aliases_created
        self.conflicts.extend(other.conflicts)


@dataclass
class BudgetImportReport:
    files: list[str] = field(default_factory=list)
    rows_read: int = 0
    rows_skipped: int = 0
    works_created: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    conflicts: list[ImportConflict] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "files": self.files,
            "rows_read": self.rows_read,
            "rows_skipped": self.rows_skipped,
            "works_created": self.works_created,
            "created": self.created,
            "updated": self.updated,
            "unchanged": self.unchanged,
            "conflicts": [conflict.__dict__ for conflict in self.conflicts],
        }

    def add(self, other: "BudgetImportReport") -> None:
        self.files.extend(other.files)
        self.rows_read += other.rows_read
        self.rows_skipped += other.rows_skipped
        self.works_created += other.works_created
        self.created += other.created
        self.updated += other.updated
        self.unchanged += other.unchanged
        self.conflicts.extend(other.conflicts)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"\s+", " ", text)
    return text.casefold()


def normalize_header(value: object) -> str:
    text = normalize_text(value)
    text = text.replace(".", "").replace("º", "o")
    return re.sub(r"[^a-z0-9/* -]", "", text).strip()


def digits_only(value: object) -> str:
    return re.sub(r"\D", "", str(value or ""))


def document_type(document: str) -> str:
    if len(document) == 11:
        return CounterpartyDocument.DocumentType.CPF
    if len(document) == 14:
        return CounterpartyDocument.DocumentType.CNPJ
    return CounterpartyDocument.DocumentType.OTHER


def import_counterparty_workbooks(
    supplier_path: str | Path | None = None,
    worker_path: str | Path | None = None,
    dry_run: bool = False,
) -> ImportReport:
    report = ImportReport()
    with transaction.atomic():
        if supplier_path:
            report.add(import_counterparty_workbook(supplier_path, Counterparty.Kind.SUPPLIER))
        if worker_path:
            report.add(import_counterparty_workbook(worker_path, Counterparty.Kind.WORKER))
        if dry_run:
            transaction.set_rollback(True)
    return report


def import_counterparty_workbook(path: str | Path, kind: str) -> ImportReport:
    path = Path(path)
    report = ImportReport(files=[str(path)])
    workbook = load_workbook(path, read_only=True, data_only=True)

    for sheet in workbook.worksheets:
        header_row_number, header = find_header(sheet)
        if not header:
            continue
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            raw = row_to_dict(header, row)
            if not any(raw.values()):
                continue
            report.rows_read += 1
            import_counterparty_row(raw, row_number, kind, report)

    return report


def find_header(sheet) -> tuple[int, dict[int, str]]:
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        header = {}
        for index, value in enumerate(row):
            normalized = normalize_header(value)
            field = HEADER_ALIASES.get(normalized)
            if field:
                header[index] = field
        if "name" in header.values():
            return row_number, header
    return 0, {}


def row_to_dict(header: dict[int, str], row: tuple) -> dict:
    date = {}
    for index, field in header.items():
        value = row[index] if index < len(row) else None
        if value is not None and str(value).strip() != "":
            date[field] = str(value).strip()
    return date


def import_counterparty_row(raw: dict, row_number: int, kind: str, report: ImportReport) -> None:
    name = clean_name(raw.get("name"))
    normalized_name = normalize_text(name)
    if not name or not normalized_name:
        report.rows_skipped += 1
        return

    documents = pick_documents(raw)
    document = pick_primary_document(raw, documents)
    person_type = pick_person_type(raw, document)
    counterparty = find_existing_counterparty(documents, normalized_name)

    if counterparty is None:
        counterparty = Counterparty.objects.create(
            name=name,
            normalized_name=normalized_name,
            kind=kind,
            person_type=person_type,
            primary_document=document,
            source=Origin.IMPORT,
            confidence=1 if document else 0,
            notes=build_notes(raw),
        )
        report.created += 1
    else:
        before_documents = report.documents_created
        before_aliases = report.aliases_created
        changed = update_counterparty(
            counterparty,
            name,
            normalized_name,
            document,
            person_type,
            kind,
            raw,
            row_number,
            report,
        )
        for related_document in documents:
            ensure_document(counterparty, related_document, report)
        ensure_alias(counterparty, name, normalized_name, report)
        legal_name = clean_name(raw.get("legal_name"))
        if legal_name:
            ensure_alias(counterparty, legal_name, normalize_text(legal_name), report)
        if changed or report.documents_created > before_documents or report.aliases_created > before_aliases:
            report.updated += 1
        else:
            report.unchanged += 1
        return

    for related_document in documents:
        ensure_document(counterparty, related_document, report)
    ensure_alias(counterparty, name, normalized_name, report)
    legal_name = clean_name(raw.get("legal_name"))
    if legal_name:
        ensure_alias(counterparty, legal_name, normalize_text(legal_name), report)


def clean_name(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def import_budget_workbooks(paths: list[str | Path], dry_run: bool = False) -> BudgetImportReport:
    report = BudgetImportReport()
    with transaction.atomic():
        for path in paths:
            report.add(import_budget_workbook(path))
        if dry_run:
            transaction.set_rollback(True)
    return report


def import_budget_workbook(path: str | Path) -> BudgetImportReport:
    path = Path(path)
    report = BudgetImportReport(files=[str(path)])
    workbook = load_workbook(path, read_only=False, data_only=True)

    for sheet in workbook.worksheets:
        work_name = read_budget_work_name(sheet) or path.stem.replace("Budget Sintético - ", "")
        work, created = Work.objects.get_or_create(
            normalized_name=normalize_text(work_name),
            defaults={"name": work_name},
        )
        if created:
            report.works_created += 1
        header_row_number, header = find_budget_header(sheet)
        if not header:
            report.conflicts.append(
                ImportConflict(
                    row_number=0,
                    name=work_name,
                    document="",
                    reason="budget_header_not_found",
                    detail=str(path),
                )
            )
            continue
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            raw = row_to_dict(header, row)
            if not any(raw.values()):
                continue
            report.rows_read += 1
            import_budget_row(work, raw, row_number, path, report)
    return report


def import_budget_workbook_for_work(path: str | Path, work: Work) -> BudgetImportReport:
    path = Path(path)
    report = BudgetImportReport(files=[str(path)])
    workbook = load_workbook(path, read_only=False, data_only=True)

    for sheet in workbook.worksheets:
        declared_work_name = read_budget_work_name(sheet)
        if declared_work_name and normalize_text(declared_work_name) != work.normalized_name:
            report.conflicts.append(
                ImportConflict(
                    row_number=0,
                    name=declared_work_name,
                    document="",
                    reason="budget_work_mismatch",
                    detail=f"The spreadsheet declares a different project than the selected project: {work.name}.",
                )
            )
            continue
        header_row_number, header = find_budget_header(sheet)
        if not header:
            report.conflicts.append(
                ImportConflict(
                    row_number=0,
                    name=work.name,
                    document="",
                    reason="budget_header_not_found",
                    detail=str(path),
                )
            )
            continue
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            raw = row_to_dict(header, row)
            if not any(raw.values()):
                continue
            report.rows_read += 1
            import_budget_row(work, raw, row_number, path, report)
    return report


def read_budget_work_name(sheet) -> str:
    for row in sheet.iter_rows(min_row=1, max_row=8, values_only=True):
        values = list(row)
        for index, value in enumerate(values):
            if normalize_header(value) in {"project", "obra"} and index + 1 < len(values):
                return clean_name(values[index + 1])
    return ""


def find_budget_header(sheet) -> tuple[int, dict[int, str]]:
    aliases = {
        "indice": "index",
        "etapa/item": "item_type",
        "etapa item": "item_type",
        "codigo": "code",
        "base": "base",
        "type": "service_type",
        "descricao": "description",
        "unid": "unit",
        "qtde": "quantity",
        "custo unitario": "unit_cost",
        "custo total": "total_cost",
    }
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        header = {}
        for index, value in enumerate(row):
            field = aliases.get(normalize_header(value))
            if field:
                header[index] = field
        if {"index", "item_type", "description"}.issubset(set(header.values())):
            return row_number, header
    return 0, {}


def import_budget_row(
    work: Work,
    raw: dict,
    row_number: int,
    path: Path,
    report: BudgetImportReport,
) -> None:
    index = clean_name(raw.get("index"))
    description = clean_name(raw.get("description"))
    if not index or not description:
        report.rows_skipped += 1
        return

    defaults = {
        "parent_index": parent_budget_index(index),
        "item_type": normalize_budget_item_type(raw.get("item_type")),
        "code": clean_name(raw.get("code")),
        "base": clean_name(raw.get("base")),
        "service_type": clean_name(raw.get("service_type")),
        "description": description,
        "normalized_description": normalize_text(description),
        "unit": clean_name(raw.get("unit")),
        "quantity": parse_optional_decimal(raw.get("quantity")),
        "unit_cost": parse_optional_decimal(raw.get("unit_cost")),
        "total_cost": parse_optional_decimal(raw.get("total_cost")),
        "source_file": str(path),
        "source_row": row_number,
        "is_active": True,
    }
    budget_item, created = BudgetItem.objects.get_or_create(work=work, index=index, defaults=defaults)
    if created:
        report.created += 1
        return

    changed = False
    for field, value in defaults.items():
        if getattr(budget_item, field) != value:
            setattr(budget_item, field, value)
            changed = True
    if changed:
        budget_item.save()
        report.updated += 1
    else:
        report.unchanged += 1


def parent_budget_index(index: str) -> str:
    parts = index.split(".")
    if len(parts) <= 1:
        return ""
    return ".".join(parts[:-1])


def normalize_budget_item_type(value: object) -> str:
    normalized = normalize_text(value)
    if normalized == "etapa":
        return BudgetItem.ItemType.STAGE
    if normalized == "subetapa":
        return BudgetItem.ItemType.SUBSTAGE
    if normalized == "item":
        return BudgetItem.ItemType.ITEM
    return BudgetItem.ItemType.OTHER


def parse_optional_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def pick_documents(raw: dict) -> list[str]:
    documents = []
    for key in ("document", "cpf", "cnpj"):
        document = digits_only(raw.get(key))
        if document and document not in documents:
            documents.append(document)
    return documents


def pick_document(raw: dict) -> str:
    return pick_primary_document(raw, pick_documents(raw))


def pick_primary_document(raw: dict, documents: list[str]) -> str:
    if not documents:
        return ""
    raw_type = normalize_text(raw.get("person_type"))
    if "juridica" in raw_type:
        cnpj = next((document for document in documents if len(document) == 14), "")
        if cnpj:
            return cnpj
    if "fisica" in raw_type:
        cpf = next((document for document in documents if len(document) == 11), "")
        if cpf:
            return cpf
    cnpj = next((document for document in documents if len(document) == 14), "")
    return cnpj or documents[0]


def pick_person_type(raw: dict, document: str) -> str:
    raw_type = normalize_text(raw.get("person_type"))
    if "juridica" in raw_type or len(document) == 14:
        return Counterparty.PersonType.COMPANY
    if "fisica" in raw_type or len(document) == 11:
        return Counterparty.PersonType.INDIVIDUAL
    return Counterparty.PersonType.UNKNOWN


def find_existing_counterparty(documents: list[str], normalized_name: str) -> Counterparty | None:
    for document in documents:
        document_record = (
            CounterpartyDocument.objects.select_related("counterparty").filter(number=document).first()
        )
        if document_record:
            return document_record.counterparty
        counterparty = Counterparty.objects.filter(primary_document=document).first()
        if counterparty:
            return counterparty
    return Counterparty.objects.filter(normalized_name=normalized_name, primary_document="").first()


def update_counterparty(
    counterparty: Counterparty,
    name: str,
    normalized_name: str,
    document: str,
    person_type: str,
    kind: str,
    raw: dict,
    row_number: int,
    report: ImportReport,
) -> bool:
    changed = False
    if counterparty.normalized_name != normalized_name:
        ensure_alias(counterparty, name, normalized_name, report)
    if should_replace_primary_document(counterparty, document, person_type):
        counterparty.primary_document = document
        changed = True
    if counterparty.person_type == Counterparty.PersonType.UNKNOWN and person_type != Counterparty.PersonType.UNKNOWN:
        counterparty.person_type = person_type
        changed = True
    if counterparty.source != Origin.IMPORT:
        counterparty.source = Origin.IMPORT
        changed = True
    if counterparty.kind != kind:
        old_kind = counterparty.kind
        counterparty.kind = kind
        changed = True
        report.conflicts.append(
            ImportConflict(
                row_number=row_number,
                name=name,
                document=document,
                reason="kind_conflict",
                detail=f"existing={old_kind}; incoming={kind}; final={counterparty.kind}",
            )
        )
    notes = build_notes(raw)
    if notes and notes not in counterparty.notes:
        counterparty.notes = f"{counterparty.notes}\n{notes}".strip()
        changed = True
    if changed:
        counterparty.save()
    return changed


def ensure_document(counterparty: Counterparty, document: str, report: ImportReport) -> None:
    if not document:
        return
    existing = CounterpartyDocument.objects.filter(number=document).first()
    if existing:
        if existing.counterparty_id == counterparty.pk and counterparty.primary_document == document and not existing.is_primary:
            CounterpartyDocument.objects.filter(counterparty=counterparty, is_primary=True).exclude(pk=existing.pk).update(
                is_primary=False
            )
            existing.is_primary = True
            existing.save(update_fields=["is_primary", "updated_at"])
        return
    is_primary = counterparty.primary_document == document
    if is_primary:
        CounterpartyDocument.objects.filter(counterparty=counterparty, is_primary=True).update(is_primary=False)
    CounterpartyDocument.objects.create(
        counterparty=counterparty,
        document_type=document_type(document),
        number=document,
        source=Origin.IMPORT,
        confidence=1,
        is_primary=is_primary,
    )
    report.documents_created += 1


def ensure_alias(counterparty: Counterparty, name: str, normalized_name: str, report: ImportReport) -> None:
    if not name or not normalized_name or normalized_name == counterparty.normalized_name:
        return
    _, created = CounterpartyAlias.objects.get_or_create(
        counterparty=counterparty,
        normalized_name=normalized_name,
        defaults={"name": name, "source": Origin.IMPORT},
    )
    if created:
        report.aliases_created += 1


def should_replace_primary_document(counterparty: Counterparty, document: str, person_type: str) -> bool:
    if not document:
        return False
    if not counterparty.primary_document:
        return True
    return (
        person_type == Counterparty.PersonType.COMPANY
        and len(document) == 14
        and len(counterparty.primary_document) == 11
    )


def build_notes(raw: dict) -> str:
    parts = []
    for label, key in (
        ("Status", "status"),
        ("Telefone", "phone"),
        ("E-mail", "email"),
        ("Bank account", "bank_account"),
        ("Cargo", "role"),
        ("Fonte", "source_note"),
    ):
        value = clean_name(raw.get(key))
        if value:
            parts.append(f"{label}: {value}")
    return "; ".join(parts)
