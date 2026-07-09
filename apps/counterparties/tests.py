from io import StringIO
from pathlib import Path
from tempfile import TemporaryDirectory

from django.contrib import admin
from django.conf import settings
from django.core.management import call_command
from django.db import IntegrityError, transaction
from django.test import TestCase
from openpyxl import Workbook, load_workbook

from .importers import (
    import_budget_workbook_for_work,
    import_budget_workbooks,
    import_counterparty_workbooks,
    normalize_text,
)
from .models import (
    BudgetImportBatch,
    BudgetItem,
    Category,
    ChartOfAccount,
    CostCenter,
    Counterparty,
    CounterpartyAlias,
    CounterpartyDocument,
    Work,
)


class CounterpartyModelTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name="Materiais", normalized_name="materiais")
        self.chart_account = ChartOfAccount.objects.create(name="Projects", normalized_name="projects")
        self.cost_center = CostCenter.objects.create(name="Sertãozinho", normalized_name="sertaozinho")
        self.work = Work.objects.create(name="Sertãozinho", normalized_name="project-sertaozinho")
        self.counterparty = Counterparty.objects.create(
            name="Vendor Teste",
            normalized_name="vendor teste",
            kind=Counterparty.Kind.SUPPLIER,
            person_type=Counterparty.PersonType.COMPANY,
            primary_document="12345678000199",
            default_category=self.category,
            default_chart_account=self.chart_account,
            default_cost_center=self.cost_center,
            default_work=self.work,
        )

    def test_core_counterparty_models_are_admin_registered(self):
        for model in (
            BudgetItem,
            Category,
            ChartOfAccount,
            CostCenter,
            Work,
            Counterparty,
            CounterpartyAlias,
            CounterpartyDocument,
            BudgetImportBatch,
        ):
            with self.subTest(model=model.__name__):
                self.assertTrue(admin.site.is_registered(model))

    def test_budget_item_is_unique_per_work_and_index(self):
        BudgetItem.objects.create(
            work=self.work,
            index="3.4",
            item_type=BudgetItem.ItemType.SUBSTAGE,
            description="CALÇADA",
            normalized_description="calcada",
        )

        with self.assertRaises(IntegrityError), transaction.atomic():
            BudgetItem.objects.create(
                work=self.work,
                index="3.4",
                item_type=BudgetItem.ItemType.SUBSTAGE,
                description="CALÇADA duplicada",
                normalized_description="calcada duplicada",
            )

    def test_category_normalized_name_must_be_unique(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            Category.objects.create(name="Materiais duplicado", normalized_name="materiais")

    def test_alias_is_unique_per_counterparty(self):
        CounterpartyAlias.objects.create(
            counterparty=self.counterparty,
            name="Vendor T.",
            normalized_name="vendor t",
        )

        with self.assertRaises(IntegrityError), transaction.atomic():
            CounterpartyAlias.objects.create(
                counterparty=self.counterparty,
                name="Vendor T",
                normalized_name="vendor t",
            )

    def test_same_alias_can_exist_for_different_counterparties(self):
        other = Counterparty.objects.create(name="Outro", normalized_name="outro")
        CounterpartyAlias.objects.create(
            counterparty=self.counterparty,
            name="Apelido",
            normalized_name="apelido",
        )
        CounterpartyAlias.objects.create(counterparty=other, name="Apelido", normalized_name="apelido")

        self.assertEqual(CounterpartyAlias.objects.filter(normalized_name="apelido").count(), 2)

    def test_document_number_must_be_unique_globally(self):
        CounterpartyDocument.objects.create(
            counterparty=self.counterparty,
            document_type=CounterpartyDocument.DocumentType.CNPJ,
            number="12345678000199",
            is_primary=True,
        )

        other = Counterparty.objects.create(name="Outro", normalized_name="outro")
        with self.assertRaises(IntegrityError), transaction.atomic():
            CounterpartyDocument.objects.create(
                counterparty=other,
                document_type=CounterpartyDocument.DocumentType.CNPJ,
                number="12345678000199",
            )

    def test_primary_document_must_be_unique_when_present(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            Counterparty.objects.create(
                name="Documento repetido",
                normalized_name="documento repetido",
                primary_document="12345678000199",
            )

    def test_blank_primary_document_can_repeat_for_pending_review_records(self):
        Counterparty.objects.create(name="Sem documento 1", normalized_name="sem documento 1")
        Counterparty.objects.create(name="Sem documento 2", normalized_name="sem documento 2")

        self.assertEqual(Counterparty.objects.filter(primary_document="").count(), 2)

    def test_counterparty_confidence_must_be_between_zero_and_one(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            Counterparty.objects.create(
                name="Confidence inválida",
                normalized_name="confianca invalida",
                confidence="1.50",
            )

    def test_only_one_primary_document_is_allowed_per_counterparty(self):
        CounterpartyDocument.objects.create(
            counterparty=self.counterparty,
            document_type=CounterpartyDocument.DocumentType.CNPJ,
            number="12345678000199",
            is_primary=True,
        )

        with self.assertRaises(IntegrityError), transaction.atomic():
            CounterpartyDocument.objects.create(
                counterparty=self.counterparty,
                document_type=CounterpartyDocument.DocumentType.CPF,
                number="12345678901",
                is_primary=True,
            )

    def test_document_confidence_must_be_between_zero_and_one(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            CounterpartyDocument.objects.create(
                counterparty=self.counterparty,
                document_type=CounterpartyDocument.DocumentType.CNPJ,
                number="98765432000188",
                confidence="1.25",
            )


class CounterpartyImporterTests(TestCase):
    def setUp(self):
        self.tempdir = TemporaryDirectory()
        self.base_path = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def test_import_is_idempotent(self):
        suppliers = self.write_workbook(
            "vendors.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [
                ["ACME Materiais", None, "12345678000199", "Active", "Legal entity"],
            ],
        )
        workers = self.write_workbook(
            "workers.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type", "Cargo"],
            [
                ["João da Silva", "11122233344", None, "Active", "Individual", "Pedreiro"],
            ],
        )

        first_report = import_counterparty_workbooks(suppliers, workers)
        second_report = import_counterparty_workbooks(suppliers, workers)

        self.assertEqual(first_report.created, 2)
        self.assertEqual(second_report.created, 0)
        self.assertEqual(second_report.updated, 0)
        self.assertEqual(second_report.unchanged, 2)
        self.assertEqual(Counterparty.objects.count(), 2)
        self.assertEqual(CounterpartyDocument.objects.count(), 2)

    def test_import_deduplicates_by_document_and_keeps_alias(self):
        suppliers = self.write_workbook(
            "vendors.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [
                ["ACME Materiais", None, "12345678000199", "Active", "Legal entity"],
                ["ACME Comércio", None, "12345678000199", "Active", "Legal entity"],
            ],
        )

        report = import_counterparty_workbooks(supplier_path=suppliers)

        counterparty = Counterparty.objects.get(primary_document="12345678000199")
        self.assertEqual(report.created, 1)
        self.assertEqual(report.updated, 1)
        self.assertEqual(Counterparty.objects.count(), 1)
        self.assertTrue(counterparty.aliases.filter(normalized_name="acme comercio").exists())

    def test_import_deduplicates_by_normalized_name_when_document_is_missing(self):
        suppliers = self.write_workbook(
            "vendors.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [
                ["Sem Documento Ltda", None, None, "Active", "Legal entity"],
                ["Sem Documento Ltda", None, None, "Active", "Legal entity"],
            ],
        )

        report = import_counterparty_workbooks(supplier_path=suppliers)

        self.assertEqual(report.created, 1)
        self.assertEqual(report.updated, 0)
        self.assertEqual(report.unchanged, 1)
        self.assertEqual(Counterparty.objects.count(), 1)

    def test_import_keeps_cpf_and_cnpj_for_company_supplier(self):
        suppliers = self.write_workbook(
            "vendors.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [
                ["Maria Jose Pontes da Silva", "05447779480", "34830383000147", "Active", "Legal entity"],
            ],
        )

        report = import_counterparty_workbooks(supplier_path=suppliers)

        counterparty = Counterparty.objects.get(normalized_name="maria jose pontes da silva")
        self.assertEqual(report.created, 1)
        self.assertEqual(report.documents_created, 2)
        self.assertEqual(counterparty.primary_document, "34830383000147")
        self.assertEqual(counterparty.person_type, Counterparty.PersonType.COMPANY)
        self.assertEqual(
            set(counterparty.documents.values_list("number", flat=True)),
            {"05447779480", "34830383000147"},
        )
        self.assertEqual(counterparty.documents.get(is_primary=True).number, "34830383000147")

    def test_worker_file_takes_priority_when_same_document_exists_as_supplier(self):
        suppliers = self.write_workbook(
            "vendors.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [
                ["Adilson Gabriel de Pontes", "53088735420", None, "Active", "Individual"],
            ],
        )
        workers = self.write_workbook(
            "workers.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type", "Cargo"],
            [
                ["Adilson Gabriel de Pontes", "53088735420", None, "Active", "Individual", "Pedreiro"],
            ],
        )

        report = import_counterparty_workbooks(suppliers, workers)

        counterparty = Counterparty.objects.get(primary_document="53088735420")
        self.assertEqual(Counterparty.objects.count(), 1)
        self.assertEqual(counterparty.kind, Counterparty.Kind.WORKER)
        self.assertEqual(len(report.conflicts), 1)
        self.assertEqual(report.conflicts[0].reason, "kind_conflict")

    def test_report_contains_created_updated_and_conflicts(self):
        suppliers = self.write_workbook(
            "vendors.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [["Pessoa Duplicada", "11122233344", None, "Active", "Individual"]],
        )
        workers = self.write_workbook(
            "workers.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [["Pessoa Duplicada", "11122233344", None, "Active", "Individual"]],
        )

        report = import_counterparty_workbooks(suppliers, workers).as_dict()

        self.assertEqual(report["created"], 1)
        self.assertEqual(report["updated"], 1)
        self.assertEqual(report["unchanged"], 0)
        self.assertEqual(report["documents_created"], 1)
        self.assertEqual(len(report["conflicts"]), 1)

    def test_dry_run_does_not_persist_records(self):
        suppliers = self.write_workbook(
            "vendors.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [["ACME Materiais", None, "12345678000199", "Active", "Legal entity"]],
        )

        report = import_counterparty_workbooks(supplier_path=suppliers, dry_run=True)

        self.assertEqual(report.created, 1)
        self.assertEqual(Counterparty.objects.count(), 0)

    def test_management_command_uses_passed_normalized_files(self):
        suppliers = self.write_workbook(
            "vendors.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [["ACME Materiais", None, "12345678000199", "Active", "Legal entity"]],
        )
        workers = self.write_workbook(
            "workers.xlsx",
            ["Name", "CPF", "CNPJ", "Status", "Type"],
            [["João da Silva", "11122233344", None, "Active", "Individual"]],
        )
        out = StringIO()

        call_command("import_counterparties", "--suppliers", suppliers, "--workers", workers, stdout=out)

        self.assertIn("Counterparty import completed.", out.getvalue())
        self.assertEqual(Counterparty.objects.count(), 2)

    def write_workbook(self, filename, headers, rows):
        path = self.base_path / filename
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dados"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        return path


class BudgetItemImporterTests(TestCase):
    def setUp(self):
        self.tempdir = TemporaryDirectory()
        self.base_path = Path(self.tempdir.name)

    def tearDown(self):
        self.tempdir.cleanup()

    def test_import_budget_items_is_idempotent(self):
        budget = self.write_budget_workbook()

        first_report = import_budget_workbooks([budget])
        second_report = import_budget_workbooks([budget])

        self.assertEqual(first_report.created, 3)
        self.assertEqual(second_report.created, 0)
        self.assertEqual(second_report.unchanged, 3)
        self.assertEqual(BudgetItem.objects.count(), 3)

    def test_import_budget_items_updates_existing_item(self):
        budget = self.write_budget_workbook(description="CALÇADA")
        import_budget_workbooks([budget])
        budget = self.write_budget_workbook(description="CALÇADA REVISADA")

        report = import_budget_workbooks([budget])

        item = BudgetItem.objects.get(index="3.4")
        self.assertEqual(report.updated, 1)
        self.assertEqual(item.description, "CALÇADA REVISADA")

    def test_import_budget_items_sets_parent_and_type(self):
        budget = self.write_budget_workbook()

        import_budget_workbooks([budget])

        item = BudgetItem.objects.get(index="3.4.6")
        self.assertEqual(item.work.name, "Sertãozinho")
        self.assertEqual(item.parent_index, "3.4")
        self.assertEqual(item.item_type, BudgetItem.ItemType.ITEM)
        self.assertEqual(item.description, "ALVENARIA DE CALÇADA")

    def test_management_command_imports_budget_items(self):
        budget = self.write_budget_workbook()
        out = StringIO()

        call_command("import_budget_items", "--path", budget, stdout=out)

        self.assertIn("Budget item import completed.", out.getvalue())
        self.assertEqual(BudgetItem.objects.count(), 3)

    def test_import_budget_items_for_existing_work_uses_selected_work(self):
        work = Work.objects.create(name="Tacima", normalized_name="tacima")
        budget = self.write_budget_workbook(work_name="Tacima")

        report = import_budget_workbook_for_work(budget, work)

        self.assertEqual(report.created, 3)
        self.assertEqual(BudgetItem.objects.filter(work=work).count(), 3)
        self.assertEqual(Work.objects.count(), 1)

    def test_import_budget_items_for_existing_work_reports_mismatch_conflict(self):
        work = Work.objects.create(name="Tacima", normalized_name="tacima")
        budget = self.write_budget_workbook(work_name="Outra Project")

        report = import_budget_workbook_for_work(budget, work)

        self.assertEqual(report.created, 0)
        self.assertEqual(BudgetItem.objects.count(), 0)
        self.assertEqual(len(report.conflicts), 1)
        self.assertEqual(report.conflicts[0].reason, "budget_work_mismatch")

    def write_budget_workbook(self, description="CALÇADA", work_name="Sertãozinho"):
        path = self.base_path / "orcamento.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Budget Sintético"
        sheet.append(["OBRA", work_name])
        sheet.append([])
        sheet.append(["ÍNDICE", "ETAPA/ITEM", "CÓDIGO", "BASE", "TIPO", "DESCRIÇÃO", "UNID.", "QTDE.", "CUSTO UNITÁRIO", "CUSTO TOTAL"])
        sheet.append(["3", "Step", "", "", "", "RUA PROJETADA 03", "", 0, 0, 1000])
        sheet.append(["3.4", "Subetapa", "", "", "", description, "", 0, 0, 500])
        sheet.append(["3.4.6", "Item", "#3.4.6", "Sem base", "Composição", "ALVENARIA DE CALÇADA", "m²", 10, 95.72, 957.2])
        workbook.save(path)
        return path


class CounterpartySourceFileIntegrityTests(TestCase):
    def test_normalized_source_files_have_no_internal_or_cross_duplicates(self):
        if not settings.SUPPLIERS_NORMALIZED_FILE.exists() or not settings.WORKERS_NORMALIZED_FILE.exists():
            self.skipTest("Normalized source workbooks are not present in this checkout.")
        suppliers = self.read_source_rows(settings.SUPPLIERS_NORMALIZED_FILE)
        workers = self.read_source_rows(settings.WORKERS_NORMALIZED_FILE)

        self.assertEqual(self.find_duplicates(suppliers), [])
        self.assertEqual(self.find_duplicates(workers), [])

        supplier_keys = {self.source_key(row) for row in suppliers}
        worker_keys = {self.source_key(row) for row in workers}
        self.assertEqual(supplier_keys & worker_keys, set())

    def read_source_rows(self, path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [self.normalize(cell.value) for cell in sheet[1]]
        indexes = {header: index for index, header in enumerate(headers)}
        rows = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            name = str(row[indexes["name"]] or "").strip()
            if not name:
                continue
            cpf = self.only_digits(row[indexes["cpf"]]) if "cpf" in indexes else ""
            cnpj = self.only_digits(row[indexes["cnpj"]]) if "cnpj" in indexes else ""
            rows.append(
                {
                    "row_number": row_number,
                    "name": name,
                    "normalized_name": self.normalize(name),
                    "document": cpf or cnpj,
                }
            )
        return rows

    def find_duplicates(self, rows):
        seen = {}
        duplicates = []
        for row in rows:
            key = self.source_key(row)
            if key in seen:
                duplicates.append((key, seen[key]["row_number"], row["row_number"]))
            seen[key] = row
        return duplicates

    def source_key(self, row):
        if row["document"]:
            return "document", row["document"]
        return "name", row["normalized_name"]

    def normalize(self, value):
        return normalize_text(value)

    def only_digits(self, value):
        return "".join(char for char in str(value or "") if char.isdigit())
